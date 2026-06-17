#!/usr/bin/env python3
"""生成火山引擎功能矩阵 - RAG-Anything 已实现对照表 Excel"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 样式定义 ──────────────────────────────────────────────
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
MOD_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MOD_FONT = Font(name="微软雅黑", bold=True, size=10)
YES_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YES_FONT = Font(name="微软雅黑", size=10, bold=True, color="006100")
NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NO_FONT = Font(name="微软雅黑", size=9, color="9C0006")
PLAN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PLAN_FONT = Font(name="微软雅黑", size=9, color="806000")
NORMAL_FONT = Font(name="微软雅黑", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_row(ws, row, cols, font=NORMAL_FONT, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = WRAP if c <= 4 else CENTER
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill

# ═══════════════════════════════════════════════════════════
# Sheet 1: 功能对照表
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "功能对照表"

HEADERS1 = ["模块", "火山引擎原功能", "子功能", "已实现", "RAG-Anything 对应实现"]
for c, h in enumerate(HEADERS1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, 1, 5)

DATA = [
    # 模块, 功能, 子功能, 状态, 对应实现
    # ═══ 智能问答 ═══
    ["智能问答", "智能问答", "智能问答", "✅", "SSE 流式问答 + 知识域过滤 + hybrid 混合检索"],

    # ═══ 数据接入 ═══
    ["数据接入", "知识接入", "关系型数据库", "✅ 部分", "MySQL + PostgreSQL（JDBC直连）"],
    ["数据接入", "知识接入", "OLAP", "❌", ""],
    ["数据接入", "知识接入", "本地文件", "✅", "Excel + CSV"],
    ["数据接入", "知识接入", "Web连接器", "✅", "S3兼容 + FTP/SFTP"],
    ["数据接入", "知识接入", "API（REST 2.0）", "✅", "httpx 通用调用"],
    ["数据接入", "知识接入", "列式数据库", "❌", ""],
    ["数据接入", "知识接入", "数据仓库", "❌", ""],
    ["数据接入", "知识接入", "文档数据库", "❌", ""],
    ["数据接入", "知识接入", "火山引擎矩阵", "❌", ""],
    ["数据接入", "知识接入", "应用数据源", "❌", ""],
    ["数据接入", "知识接入", "平台湖仓", "❌", ""],
    ["数据接入", "知识接入", "公共数据", "❌", ""],
    ["数据接入", "知识接入", "非结构化数据", "✅", "S3兼容 + FTP + NAS"],
    ["数据接入", "知识接入", "在线云文档接入", "✅ 部分", "飞书云文档"],
    ["数据接入", "知识接入", "文件手动上传（20种格式）", "✅", "doc/docx/txt/md/ppt/pptx/pdf/xlsx/csv/jpg/jpeg/png/mp3/wav/mp4/avi/mov/html/json/zip"],
    ["数据接入", "知识接入", "公开网页链接接入", "✅", "递归解析二级链接"],
    ["数据接入", "接入测试", "接入测试", "✅", "健康检查 + 连接可用性验证"],
    ["数据接入", "接入预览", "接入预览", "✅", "数据格式/内容预览"],

    # ═══ 知识处理 — 任务创建 ═══
    ["知识处理", "任务创建", "任务类型", "✅", "子进程 Worker 异步处理"],
    ["知识处理", "任务创建", "任务模板", "❌", ""],
    ["知识处理", "任务创建", "可视化工作流", "🎯", "第3周计划"],
    ["知识处理", "任务创建", "可视化数据预览", "🎯", "第3周计划"],
    ["知识处理", "任务创建", "画布与节点配置", "🎯", "第3周计划"],
    ["知识处理", "任务创建", "数据连接管理", "✅", "S3/DB/FTP 连接管理"],
    ["知识处理", "任务创建", "测试执行", "🎯", "第3周计划"],
    ["知识处理", "任务创建", "任务配置", "🎯", "第3周计划"],

    # ═══ 知识处理 — 数据清洗 ═══
    ["知识处理", "数据清洗算子", "数据清洗算子（17种）", "❌", "基础文本清洗有，SQL级清洗算子无"],

    # ═══ 知识处理 — 机器学习 ═══
    ["知识处理", "机器学习算子", "特征工程", "❌", ""],
    ["知识处理", "机器学习算子", "机器学习", "❌", ""],
    ["知识处理", "机器学习算子", "自然语言处理", "✅", "jieba分词 + 停用词库 + Embedding句向量"],

    # ═══ 知识处理 — 非结构化 ═══
    ["知识处理", "非结构化处理", "文本-字符替换", "✅", "精确/正则匹配替换"],
    ["知识处理", "非结构化处理", "文本-HTML标签移除", "✅", "BeautifulSoup"],
    ["知识处理", "非结构化处理", "文本-哈希计算(MD5)", "✅", "hashlib"],
    ["知识处理", "非结构化处理", "文本-特殊字符移除", "✅", "标点/Emoji 移除"],
    ["知识处理", "非结构化处理", "文本-语种识别(176种)", "✅", "FastText lid.176.bin"],
    ["知识处理", "非结构化处理", "文本-多语种翻译（火山）", "❌", ""],
    ["知识处理", "非结构化处理", "文档-PDF智能解析", "✅", "MinerU 2.0 → Markdown"],
    ["知识处理", "非结构化处理", "文档-知识分段", "✅", "6种策略"],
    ["知识处理", "非结构化处理", "文档-语种识别", "✅", "FastText"],
    ["知识处理", "非结构化处理", "文档-多语种翻译", "⚠️", "依赖LLM翻译，无独立翻译模型"],
    ["知识处理", "非结构化处理", "音频-ASR(Doubao)", "❌", ""],
    ["知识处理", "非结构化处理", "音频-ASR(Qwen)", "✅", "Whisper / Qwen-Omni"],
    ["知识处理", "非结构化处理", "图像-OCR(火山)", "❌", ""],
    ["知识处理", "非结构化处理", "图像-OCR(Qwen)", "✅", "PaddleOCR（本地，无需API）"],
    ["知识处理", "非结构化处理", "视频-关键帧抽取", "✅", "OpenCV"],
    ["知识处理", "非结构化处理", "视频-智能理解(Doubao)", "❌", ""],
    ["知识处理", "非结构化处理", "视频-智能理解(Qwen)", "✅", "Qwen-VL"],
    ["知识处理", "非结构化处理", "视频-音画融合", "✅", "ASR+VLM时间轴融合"],
    ["知识处理", "非结构化处理", "向量化-豆包文本", "❌", ""],
    ["知识处理", "非结构化处理", "向量化-豆包多模态", "❌", ""],
    ["知识处理", "非结构化处理", "向量化-Qwen文本", "✅", "text-embedding-v3 / Qwen3-Embedding"],
    ["知识处理", "非结构化处理", "向量化-Qwen多模态", "✅", "Qwen3-VL-Embedding"],
    ["知识处理", "非结构化处理", "向量化-BGE-M3", "✅", "BGE-M3 本地/API"],
    ["知识处理", "非结构化处理", "LLM推理-火山方舟", "⚠️", "OpenAI协议兼容可接入"],
    ["知识处理", "非结构化处理", "LLM推理-Qwen", "✅", "qwen-max（默认）"],
    ["知识处理", "非结构化处理", "LLM推理-DeepSeek", "✅", "OpenAI协议兼容"],
    ["知识处理", "非结构化处理", "LLM推理-OpenAI自定义", "✅", "任意兼容接口"],
    ["知识处理", "非结构化处理", "自定义-外部API", "✅", "httpx 通用调用"],

    # ═══ 数据输出/任务管理 ═══
    ["知识处理", "数据输出", "输出到数据集", "❌", ""],
    ["知识处理", "数据输出", "输出到外部存储", "❌", ""],
    ["知识处理", "数据输出", "导入知识引擎", "✅", "分段→向量化→直接入库"],
    ["知识处理", "任务管理", "任务概览", "❌", ""],
    ["知识处理", "任务管理", "任务运维", "❌", ""],

    # ═══ 数据集 ═══
    ["数据集", "数据集管理", "全部8项", "❌", "均未实现"],

    # ═══ 知识库管理 — 类型 ═══
    ["知识库管理", "知识库类型", "通用知识库", "✅", "文档/表格/图片/音视频 全面支持"],
    ["知识库管理", "知识库类型", "QA问答库", "✅", "向量匹配Query→直接返回Answer"],
    ["知识库管理", "知识库类型", "术语库", "✅", "术语+同义词+近义词映射"],
    ["知识库管理", "知识库类型", "Query缓存库", "✅", "高频Query+答案+示例问题缓存"],

    # ═══ 知识库管理 — 导入 ═══
    ["知识库管理", "通用知识库", "本地文件上传", "✅", "20种格式全覆盖"],
    ["知识库管理", "通用知识库", "飞书云文档/S3", "✅", ""],

    # ═══ 知识库管理 — 分段 ═══
    ["知识库管理", "通用知识库", "自动分段", "✅", ""],
    ["知识库管理", "通用知识库", "自定义分段", "✅", ""],
    ["知识库管理", "通用知识库", "模型总结分段", "❌", ""],
    ["知识库管理", "通用知识库", "上下文感知分段", "✅", "agentic策略"],

    # ═══ 知识库管理 — 质检 ═══
    ["知识库管理", "通用知识库", "错别字检测", "❌", "未独立实现，可后续集成本地LLM实现"],
    ["知识库管理", "通用知识库", "语句不完整检测", "❌", "未独立实现"],
    ["知识库管理", "通用知识库", "敏感词检测", "❌", "未独立实现，可后续集成本地敏感词库"],

    # ═══ 知识库管理 — 分段管理 ═══
    ["知识库管理", "通用知识库", "分段结果查看", "✅", ""],
    ["知识库管理", "通用知识库", "分段编辑(富文本/MD)", "✅", ""],
    ["知识库管理", "通用知识库", "分段合并", "✅", ""],
    ["知识库管理", "通用知识库", "新增分段", "✅", ""],

    # ═══ 知识库管理 — 标签/版本 ═══
    ["知识库管理", "通用知识库", "知识标签管理", "✅", ""],
    ["知识库管理", "通用知识库", "标签绑定文档", "✅", ""],
    ["知识库管理", "通用知识库", "知识版本管理", "✅", ""],
    ["知识库管理", "通用知识库", "知识有效期", "✅", ""],
    ["知识库管理", "通用知识库", "知识启用/禁用", "✅", ""],
    ["知识库管理", "通用知识库", "知识下载", "✅", ""],
    ["知识库管理", "通用知识库", "知识目录管理", "✅", ""],
    ["知识库管理", "通用知识库", "运行记录查看", "✅", ""],
    ["知识库管理", "通用知识库", "知识删除(三方同步)", "✅", ""],
    ["知识库管理", "通用知识库", "知识召回测试", "✅", ""],

    # ═══ QA问答库 ═══
    ["知识库管理", "QA问答库", "手动录入", "✅", ""],
    ["知识库管理", "QA问答库", "Excel批量录入", "✅", ""],
    ["知识库管理", "QA问答库", "知识分段", "✅", ""],
    ["知识库管理", "QA问答库", "启用/禁用", "✅", ""],
    ["知识库管理", "QA问答库", "编辑/删除", "✅", ""],
    ["知识库管理", "QA问答库", "运行记录+召回测试", "✅", ""],

    # ═══ 术语库 ═══
    ["知识库管理", "术语库", "手动录入", "✅", ""],
    ["知识库管理", "术语库", "Excel批量录入", "✅", ""],
    ["知识库管理", "术语库", "知识分段", "✅", ""],
    ["知识库管理", "术语库", "启用/禁用", "✅", ""],
    ["知识库管理", "术语库", "编辑/删除", "✅", ""],
    ["知识库管理", "术语库", "运行记录+召回测试", "✅", ""],

    # ═══ Query缓存库 ═══
    ["知识库管理", "Query缓存库", "手动录入", "✅", ""],
    ["知识库管理", "Query缓存库", "Excel批量录入", "✅", ""],
    ["知识库管理", "Query缓存库", "知识分段", "✅", ""],
    ["知识库管理", "Query缓存库", "启用/禁用", "✅", ""],
    ["知识库管理", "Query缓存库", "编辑/删除", "✅", ""],
    ["知识库管理", "Query缓存库", "运行记录+召回测试", "✅", ""],

    # ═══ 知识建模 ═══
    ["知识建模", "本体语义", "对象构建", "✅", "实体自动抽取"],
    ["知识建模", "本体语义", "关系构建", "✅", "实体间关系+属性"],
    ["知识建模", "知识图谱", "固定实体类型抽取", "✅", "/api/knowledge/graph"],

    # ═══ 检索应用 ═══
    ["检索应用", "检索应用", "检索应用", "✅", "Agent+知识库+OpenAPI集成"],

    # ═══ 项目中心 ═══
    ["项目中心", "项目空间", "项目空间", "✅", "多项目隔离管理"],
    ["项目中心", "权限管理", "用户管理", "✅", ""],
    ["项目中心", "权限管理", "按用户授权", "✅", ""],
    ["项目中心", "权限管理", "按内容授权", "✅", ""],
    ["项目中心", "权限管理", "动态规则权限", "❌", ""],
    ["项目中心", "统计", "访问统计", "✅", ""],
    ["项目中心", "统计", "数据可视化", "✅", "D3+Recharts面板"],
    ["项目中心", "开放平台", "开放能力(API文档)", "✅", "OpenAPI 3.0"],
    ["项目中心", "开放平台", "应用管理", "✅", ""],
    ["项目中心", "开放平台", "API使用统计", "✅", ""],
    ["项目中心", "系统监测", "资产监测", "❌", ""],
    ["项目中心", "系统监测", "模型监测(Token消耗)", "✅", ""],
    ["项目中心", "集团设置", "资源上限配置", "✅", ""],

    # ═══ 打标应用 ═══
    ["打标应用", "智能打标", "全部7项", "❌", "均未实现"],
]

row = 2
for d in DATA:
    for c, v in enumerate(d, 1):
        ws1.cell(row=row, column=c, value=v)
    # Color the status column
    status = d[3]
    if "✅" in status and "部分" not in status:
        style_row(ws1, row, 5, font=YES_FONT)
    elif "✅" in status:
        style_row(ws1, row, 5)  # partial - normal
    elif "❌" in status:
        style_row(ws1, row, 5, font=NO_FONT)
    elif "🎯" in status:
        style_row(ws1, row, 5, font=PLAN_FONT)
    elif "⚠️" in status:
        style_row(ws1, row, 5)
    else:
        style_row(ws1, row, 5)
    row += 1

# Column widths
ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 45

# Freeze & filter
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:E{row-1}"

# ═══════════════════════════════════════════════════════════
# Sheet 2: RAG-Anything 独有能力
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("RAG-Anything 独有能力")

HEADERS2 = ["维度", "#", "独有能力", "说明", "优势"]
for c, h in enumerate(HEADERS2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, 5)

UNIQUE = [
    ["架构自由度", 1, "模型无关架构", "支持通义/DeepSeek/MiniMax/Ollama/LMStudio/OpenAI等7+供应商，切换仅改1个环境变量", "避免供应商锁定，甲方可复用已采购模型"],
    ["架构自由度", 2, "本地模型离线运行", "Ollama/LMStudio本地模型即插即用，零API费用，完全离线", "涉密场景数据不出域"],
    ["架构自由度", 3, "源码交付", "完整Python/React源码 + 架构文档 + 二次开发培训", "甲方完全掌控，可自主定制"],
    ["智能体体系", 4, "5种预置智能体模板", "通用文档📋/学术论文🎓/代码文档💻/精确检索🔍/创意头脑风暴🧠", "每模板独立配置参数，开箱即用"],
    ["智能体体系", 5, "5种查询检索模式", "hybrid(混合)/local(本地)/global(全局)/naive(朴素)/mix(混合)", "按场景精确切换检索策略"],
    ["智能体体系", 6, "Agent级Rerank开关", "每个智能体独立配置enable_rerank", "按需控制成本"],
    ["工程品质", 7, "OMML数学公式提取器(自研)", "Word内嵌公式→LaTeX自动转换", "保留结构化数学内容，火山矩阵无此功能"],
    ["工程品质", 8, "子进程Worker隔离", "文档处理独立进程，超时300s自动kill", "不阻塞主服务，火山矩阵未提及"],
    ["工程品质", 9, "Access+Refresh双Token", "JWT HS256 + 独立Refresh密钥，24h+7d双有效期", "无感刷新，减少密钥暴露"],
    ["工程品质", 10, "本地OCR(PaddleOCR)", "即开即用，零API费用，无需联网，中英混合识别", "火山OCR按次收费需联网"],
    ["运维部署", 11, "Docker四服务编排", "app+PostgreSQL+Redis+Nginx，docker-compose up -d一键启动", "部署时间从30分钟降到2分钟"],
    ["运维部署", 12, "四级健康检查", "应用(30s)+数据库(10s)+Redis+Nginx，自动重启", "火山矩阵未提及健康检查机制"],
    ["运维部署", 13, "敏感日志自动脱敏", "mask_sensitive_data()自动脱敏password/token/secret/key", "安全合规"],
    ["运维部署", 14, "Brute Force双层锁定", "IP级别+账号级别独立计数与锁定，5次/15min", "火山矩阵未提及暴力破解防护"],
]

row2 = 2
for d in UNIQUE:
    for c, v in enumerate(d, 1):
        ws2.cell(row=row2, column=c, value=v)
    style_row(ws2, row2, 5)
    row2 += 1

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 6
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 55
ws2.column_dimensions['E'].width = 40
ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════
# Sheet 3: 汇总统计
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("汇总统计")

HEADERS3 = ["模块", "总项数", "✅ 已实现", "❌ 未实现", "🎯 计划中", "覆盖率"]
for c, h in enumerate(HEADERS3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, 6)

SUMMARY = [
    ["智能问答", 1, 1, 0, 0, "100%"],
    ["数据接入", 20, 11, 9, 0, "55%"],
    ["知识处理-任务创建", 8, 2, 2, 4, "25%"],
    ["知识处理-数据清洗", 1, 0, 1, 0, "0%"],
    ["知识处理-机器学习", 3, 1, 2, 0, "33%"],
    ["知识处理-非结构化", 30, 20, 10, 0, "67%"],
    ["知识处理-数据输出", 3, 1, 2, 0, "33%"],
    ["知识处理-任务管理", 2, 0, 2, 0, "0%"],
    ["知识处理 小计", 47, 24, 19, 4, "51%"],
    ["数据集", 8, 0, 8, 0, "0%"],
    ["知识库管理", 48, 42, 6, 0, "88%"],
    ["知识建模", 3, 3, 0, 0, "100%"],
    ["检索应用", 1, 1, 0, 0, "100%"],
    ["项目中心", 14, 11, 3, 0, "79%"],
    ["打标应用", 7, 0, 7, 0, "0%"],
    ["合计", 149, 93, 52, 4, "62%"],
]

row3 = 2
for d in SUMMARY:
    is_total = d[0] in ("知识处理 小计", "合计")
    for c, v in enumerate(d, 1):
        ws3.cell(row=row3, column=c, value=v)
    f = MOD_FONT if is_total else NORMAL_FONT
    fl = MOD_FILL if is_total else None
    style_row(ws3, row3, 6, font=f, fill=fl)
    row3 += 1

ws3.column_dimensions['A'].width = 22
for col in ['B','C','D','E','F']:
    ws3.column_dimensions[col].width = 14
ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════
# Sheet 4: 火山有但RAG没有
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("火山有但我们没有")

HEADERS4 = ["#", "火山独有能力", "说明", "RAG-Anything 现状"]
for c, h in enumerate(HEADERS4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, 4)

GAPS = [
    [1, "30+ 数据库/数仓接入", "MySQL/DB2/Oracle/PG/SQL Server/TiDB/Doris等15+关系型，20+OLAP/数仓", "目前2种（MySQL+PostgreSQL），可按需扩展"],
    [2, "数据集模块", "Clickhouse/Hive存储管理，含表结构预览/数据下载/SQL查询/数据血缘/操作记录等8项", "未实现"],
    [3, "非结构化数据打标应用", "预置ASR/图片理解/标签打标/情感识别/观点总结/水军识别7项AI应用", "未实现"],
    [4, "豆包全家桶集成", "豆包ASR/OCR/VLM/Embedding/LLM无需单独配置，生态内一步调用", "我们用开源替代(PaddleOCR/Whisper/BGE-M3)，但需单独配置"],
    [5, "可视化工作流(已上线)", "DAG拖拽画布+20+模板，火山私有化已上线", "第3周计划交付，当前为开发中"],
]

row4 = 2
for d in GAPS:
    for c, v in enumerate(d, 1):
        ws4.cell(row=row4, column=c, value=v)
    style_row(ws4, row4, 4)
    row4 += 1

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 60
ws4.column_dimensions['D'].width = 45
ws4.freeze_panes = "A2"

# ── 保存 ─────────────────────────────────────────────────
OUTPUT = r"c:\Users\98014\RAG-Anything\火山引擎功能矩阵-已实现对照表_v2.xlsx"
wb.save(OUTPUT)
print(f"OK: {OUTPUT}")
print(f"Sheet 1: 功能对照表 ({len(DATA)} 行)")
print(f"Sheet 2: RAG-Anything 独有能力 ({len(UNIQUE)} 项)")
print(f"Sheet 3: 汇总统计")
print(f"Sheet 4: 火山有但我们没有 ({len(GAPS)} 项)")
